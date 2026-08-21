"""Gera os dois Excel do fechamento mensal, no mesmo layout de colunas visto
em 'Importações em Andamento' (aba Base) / 'Processos Fechados' (planilhas
reais da equipe):

- Razão Atualizado: todas as linhas do mês já "explodidas" por processo
  participante (ver GerarLinhasRazaoRateadoUseCase), cada uma marcada com o
  status de fechamento calculado.
- Processos Fechados: uma aba por processo com status Fechado, com as linhas,
  a linha de totais e o "Saldo processo" - a mesma "memória de cálculo" que a
  equipe hoje monta copiando e colando manualmente.
"""

import io
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font

from app.domain.entities import LinhaRazaoRateada

CABECALHO = [
    "Empresa",
    "Data",
    "Conta",
    "Numero Contabil",
    "Unidade",
    "Historico",
    "Debito",
    "Credito",
    "Movimentação",
    "Processo",
    "Processo Full",
    "Processo (Controle de Importação)",
    "Status",
]


def _linha_para_row(linha: LinhaRazaoRateada) -> list:
    return [
        linha.empresa,
        linha.data,
        linha.conta,
        linha.numero_contabil,
        linha.unidade,
        linha.historico,
        float(linha.debito),
        float(linha.credito),
        float(linha.movimentacao),
        linha.processo,
        linha.processo_full,
        linha.processo_controle_importacao,
        linha.status,
    ]


def _escrever_cabecalho(ws) -> None:
    ws.append(CABECALHO)
    for celula in ws[1]:
        celula.font = Font(bold=True)


def _nome_aba_unico(wb, codigo: str) -> str:
    base = codigo[:31]
    nome = base
    sufixo = 1
    while nome in wb.sheetnames:
        sufixo += 1
        nome = f"{base[: 31 - len(str(sufixo)) - 1]}_{sufixo}"
    return nome


def construir_razao_atualizado(linhas: list[LinhaRazaoRateada]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Razão Atualizado"
    _escrever_cabecalho(ws)
    for linha in linhas:
        ws.append(_linha_para_row(linha))
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def construir_processos_fechados(linhas_por_processo: dict[str, list[LinhaRazaoRateada]]) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for codigo in sorted(linhas_por_processo):
        linhas = linhas_por_processo[codigo]
        ws = wb.create_sheet(title=_nome_aba_unico(wb, codigo))
        _escrever_cabecalho(ws)

        soma_debito = Decimal("0")
        soma_credito = Decimal("0")
        for linha in linhas:
            ws.append(_linha_para_row(linha))
            soma_debito += linha.debito
            soma_credito += linha.credito

        ws.append([None, None, None, None, None, None, float(soma_debito), float(soma_credito)])
        ws.append([])
        ws.append(
            [None, None, None, None, None, "Saldo processo", None, None, float(soma_debito - soma_credito)]
        )

    if not wb.sheetnames:
        wb.create_sheet(title="Nenhum processo fechado")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
