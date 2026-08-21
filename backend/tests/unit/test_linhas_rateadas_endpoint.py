"""Teste de integração ponta a ponta (HTTP real, SQLite em memória) das
linhas rateadas em JSON e dos dois exports .xlsx (Razão Atualizado /
Processos Fechados) - reaproveita o mesmo padrão dos outros testes de
endpoint (sem Drive/Postgres reais)."""

import io
from datetime import date
from decimal import Decimal

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.domain import entities
from app.infrastructure.db.base import Base
from app.infrastructure.db.session import get_db
from app.infrastructure.repositories.processo_repository import SqlAlchemyProcessoRepository
from app.infrastructure.repositories.razao_repository import SqlAlchemyRazaoRepository

MES = date(2026, 6, 1)


def _client_e_sessionmaker():
    from fastapi.testclient import TestClient

    import app.main as main_module

    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(engine)
    TestingSessionLocal = sessionmaker(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    main_module.app.dependency_overrides[get_db] = override_get_db
    return TestClient(main_module.app), TestingSessionLocal


def _login(client) -> str:
    from app.core.config import get_settings

    settings = get_settings()
    resp = client.post("/auth/login", json={"email": settings.admin_email, "password": settings.admin_password})
    assert resp.status_code == 200, resp.text
    return resp.json()["access_token"]


def _semear_processo_completo(SessionLocal, codigo: str) -> None:
    db = SessionLocal()
    try:
        embarque = entities.Embarque(codigo=f"{codigo}.1", drive_folder_id="f")
        for tipo in (entities.TipoDocumento.DI, entities.TipoDocumento.INVOICE_CI, entities.TipoDocumento.NOTA_FISCAL):
            embarque.documentos.append(
                entities.Documento(
                    ref=entities.DocumentoRef(drive_file_id=f"{codigo}-{tipo}", nome_arquivo="x.pdf", caminho="x"),
                    tipo=tipo,
                )
            )
        processo = entities.Processo(codigo=codigo, empresa_codigo=codigo[:3])
        processo.embarques = [embarque]
        SqlAlchemyProcessoRepository(db).salvar(processo)
    finally:
        db.close()


def _lancamento(processo_codigo: str, debito="100", credito="0", data=None) -> entities.LancamentoRazao:
    return entities.LancamentoRazao(
        mes_referencia=MES,
        data=data or date(2026, 6, 15),
        historico=f"FRETE {processo_codigo}",
        valor_debito=Decimal(debito),
        valor_credito=Decimal(credito),
        empresa="BB",
        conta_contabil="113103",
        numero_contabil="NR: 000001",
        unidade="50001",
        processos_codigos=[processo_codigo],
        categoria_classificada=entities.CategoriaLancamento.FRETE,
    )


def _semear_lancamentos(SessionLocal, *lancamentos: entities.LancamentoRazao) -> None:
    # salvar_lote substitui o lote inteiro do mês a cada chamada (ver
    # SqlAlchemyRazaoRepository) - por isso todos os lançamentos do mês
    # precisam ir numa única chamada, nunca uma por processo.
    db = SessionLocal()
    try:
        SqlAlchemyRazaoRepository(db).salvar_lote(list(lancamentos))
    finally:
        db.close()


def test_linhas_rateadas_do_processo_via_http():
    client, SessionLocal = _client_e_sessionmaker()
    token = _login(client)
    headers = {"Authorization": f"Bearer {token}"}

    _semear_processo_completo(SessionLocal, "GOC25129")
    _semear_lancamentos(SessionLocal, _lancamento("GOC25129", debito="0", credito="0"))

    client.post("/fechamento/processar", params={"mes_referencia": "2026-06-01"}, headers=headers)

    resp = client.get(
        "/processos/GOC25129/linhas-rateadas", params={"mes_referencia": "2026-06-01"}, headers=headers
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert len(body["linhas"]) == 1
    linha = body["linhas"][0]
    assert linha["processo_full"] == "GOC25129"
    assert linha["status"] == "Fechado"
    assert linha["numero_contabil"] == "NR: 000001"
    assert linha["unidade"] == "50001"


def test_exportar_razao_atualizado_xlsx_via_http():
    client, SessionLocal = _client_e_sessionmaker()
    token = _login(client)
    headers = {"Authorization": f"Bearer {token}"}

    _semear_processo_completo(SessionLocal, "GOC25129")
    _semear_lancamentos(SessionLocal, _lancamento("GOC25129", debito="100"))
    client.post("/fechamento/processar", params={"mes_referencia": "2026-06-01"}, headers=headers)

    resp = client.get(
        "/fechamento/exportar/razao-atualizado.xlsx", params={"mes_referencia": "2026-06-01"}, headers=headers
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith("application/vnd.openxmlformats")
    assert "attachment" in resp.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Razão Atualizado"]
    linhas = list(ws.iter_rows(values_only=True))
    assert linhas[0][0] == "Empresa"
    assert len(linhas) == 2  # cabeçalho + 1 linha


def test_exportar_processos_fechados_xlsx_via_http_so_inclui_fechados():
    client, SessionLocal = _client_e_sessionmaker()
    token = _login(client)
    headers = {"Authorization": f"Bearer {token}"}

    _semear_processo_completo(SessionLocal, "GOC25129")  # vai fechar (saldo zero)
    # BBI25167 nunca sincronizado no Drive -> Bloqueado, não deve aparecer no export
    _semear_lancamentos(
        SessionLocal,
        _lancamento("GOC25129", debito="0", credito="0"),
        _lancamento("BBI25167", debito="50"),
    )

    client.post("/fechamento/processar", params={"mes_referencia": "2026-06-01"}, headers=headers)

    resp = client.get(
        "/fechamento/exportar/processos-fechados.xlsx", params={"mes_referencia": "2026-06-01"}, headers=headers
    )
    assert resp.status_code == 200, resp.text
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["GOC25129"]


def test_endpoints_de_export_exigem_autenticacao():
    client, _ = _client_e_sessionmaker()
    resp = client.get("/fechamento/exportar/razao-atualizado.xlsx", params={"mes_referencia": "2026-06-01"})
    assert resp.status_code == 401
    resp = client.get("/fechamento/exportar/processos-fechados.xlsx", params={"mes_referencia": "2026-06-01"})
    assert resp.status_code == 401
