import io
from datetime import date
from decimal import Decimal

import openpyxl

from app.domain.entities import LinhaRazaoRateada
from app.infrastructure.xlsx.exportador import CABECALHO, construir_processos_fechados, construir_razao_atualizado


def _linha(processo, processo_full, debito="100", credito="0", status="Fechado"):
    return LinhaRazaoRateada(
        empresa="BB",
        data=date(2026, 6, 15),
        conta="113103",
        numero_contabil="NR: 000001",
        unidade="50001",
        historico="PGTO TESTE",
        debito=Decimal(debito),
        credito=Decimal(credito),
        processo=processo,
        processo_full=processo_full,
        status=status,
    )


def test_razao_atualizado_tem_cabecalho_e_uma_linha_por_registro():
    linhas = [_linha("GOC25129.1", "GOC25129"), _linha("BBI25167.1", "BBI25167", debito="0", credito="50")]
    conteudo = construir_razao_atualizado(linhas)

    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb["Razão Atualizado"]
    linhas_lidas = list(ws.iter_rows(values_only=True))

    assert linhas_lidas[0] == tuple(CABECALHO)
    assert len(linhas_lidas) == 3  # cabeçalho + 2 linhas
    assert linhas_lidas[1][9] == "GOC25129.1"  # coluna Processo
    assert linhas_lidas[1][11] == "GOC25129-1"  # coluna Processo (Controle de Importação)
    assert linhas_lidas[2][6] == 0  # Debito
    assert linhas_lidas[2][7] == 50  # Credito


def test_processos_fechados_uma_aba_por_processo_com_totais_e_saldo():
    linhas_por_processo = {
        "GOC25129": [_linha("GOC25129.1", "GOC25129", debito="100"), _linha("GOC25129.2", "GOC25129", debito="50")],
        "BBI25167": [_linha("BBI25167.1", "BBI25167", debito="0", credito="30")],
    }
    conteudo = construir_processos_fechados(linhas_por_processo)

    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    assert set(wb.sheetnames) == {"GOC25129", "BBI25167"}

    ws = wb["GOC25129"]
    linhas_lidas = list(ws.iter_rows(values_only=True))
    assert linhas_lidas[0] == tuple(CABECALHO)
    assert linhas_lidas[1][9] == "GOC25129.1"
    assert linhas_lidas[2][9] == "GOC25129.2"
    # linha de totais: soma dos débitos na coluna 6 (Debito)
    assert linhas_lidas[3][6] == 150
    # linha em branco
    assert all(v is None for v in linhas_lidas[4])
    # linha "Saldo processo"
    assert linhas_lidas[5][5] == "Saldo processo"
    assert linhas_lidas[5][8] == 150


def test_processos_fechados_vazio_ainda_gera_workbook_valido():
    conteudo = construir_processos_fechados({})
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    assert len(wb.sheetnames) == 1


def test_nomes_de_aba_truncados_e_unicos_quando_colidem():
    codigo_longo_a = "PROCESSO_MUITO_LONGO_QUE_PASSA_DE_31_A"
    codigo_longo_b = "PROCESSO_MUITO_LONGO_QUE_PASSA_DE_31_B"
    linhas_por_processo = {
        codigo_longo_a: [_linha("X", codigo_longo_a)],
        codigo_longo_b: [_linha("Y", codigo_longo_b)],
    }
    conteudo = construir_processos_fechados(linhas_por_processo)
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    assert len(wb.sheetnames) == 2
    assert len(set(wb.sheetnames)) == 2
    assert all(len(nome) <= 31 for nome in wb.sheetnames)
